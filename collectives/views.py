from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
import requests
from requests.auth import HTTPBasicAuth
from datetime import datetime
import tempfile
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from .models import CollectivesConfig, PageStatus
from .utils import parse_webdav_response, build_collectives_url

STATUS_OPTIONS = ['', '📝 In Progress', '🚀 Ready to Post', '📤 Posted', '❌ Rejected']
TYP_OPTIONS = ['', 'Post', 'Other']

@login_required
def dashboard(request):
    config = CollectivesConfig.get_config()
    return render(request, 'collectives/dashboard.html', {'config': config})

@login_required
@require_http_methods(["GET", "POST"])
def api_config(request):
    config = CollectivesConfig.get_config()
    if request.method == 'POST':
        data = json.loads(request.body)
        config.nextcloud_url = data.get('nextcloud_url', '').rstrip('/')
        config.kollektive_name = data.get('kollektive_name', '')
        config.username = data.get('username', '')
        if data.get('app_password'):
            config.app_password = data.get('app_password')
        config.connected = bool(config.nextcloud_url and config.username and config.app_password)
        config.save()
        return JsonResponse({'success': True, 'config': {
            'nextcloud_url': config.nextcloud_url,
            'kollektive_name': config.kollektive_name,
            'username': config.username,
            'connected': config.connected
        }})
    return JsonResponse({
        'nextcloud_url': config.nextcloud_url,
        'kollektive_name': config.kollektive_name,
        'username': config.username,
        'connected': config.connected
    })

@login_required
@require_http_methods(["POST"])
def test_connection(request):
    config = CollectivesConfig.get_config()
    if not config.nextcloud_url or not config.username or not config.app_password:
        return JsonResponse({'success': False, 'message': 'Please fill in all fields'})
    try:
        url = f"{config.nextcloud_url}/remote.php/dav/files/{config.username}/"
        r = requests.request('PROPFIND', url,
            auth=HTTPBasicAuth(config.username, config.app_password),
            headers={'Depth': '0'}, timeout=10)
        if r.status_code in [200, 207]:
            config.connected = True
            config.save()
            return JsonResponse({'success': True, 'message': f'Connection to {config.nextcloud_url} successful!'})
        else:
            return JsonResponse({'success': False, 'message': f'Error: HTTP {r.status_code}'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Connection error: {str(e)}'})

@login_required
@require_http_methods(["GET"])
def get_status(request):
    statuses = {}
    for ps in PageStatus.objects.all():
        statuses[ps.path] = {
            'status': ps.status,
            'typ': ps.typ,
            'planned_date': ps.planned_date.isoformat() if ps.planned_date else ''
        }
    return JsonResponse(statuses)

@login_required
@csrf_exempt
@require_http_methods(["POST"])
def set_status(request):
    data = json.loads(request.body)
    path = data.get('path') or data.get('folder')
    if not path:
        return JsonResponse({'success': False, 'message': 'No path specified'})
    page_status, created = PageStatus.objects.get_or_create(path=path)
    if 'status' in data:
        page_status.status = data['status']
    if 'typ' in data:
        page_status.typ = data['typ']
    if 'planned_date' in data:
        page_status.planned_date = data['planned_date'] or None
    page_status.save()
    return JsonResponse({'success': True})

@login_required
def get_pages(request):
    config = CollectivesConfig.get_config()
    if not config.connected:
        return JsonResponse({'success': False, 'message': 'Not connected'})
    if not config.kollektive_name:
        return JsonResponse({'success': False, 'message': 'No collective configured'})
    try:
        webdav_url = f"{config.nextcloud_url}/remote.php/dav/files/{config.username}/Kollektive/{config.kollektive_name}/"
        r = requests.request('PROPFIND', webdav_url,
            auth=HTTPBasicAuth(config.username, config.app_password),
            headers={'Depth': 'infinity'}, timeout=30)
        if r.status_code in [200, 207]:
            pages = parse_webdav_response(r.text, config.username, config.kollektive_name)
            status_map = {ps.path: ps for ps in PageStatus.objects.all()}
            for page in pages:
                if page.get('is_readme'):
                    if page['folder_parts']:
                        page['url'] = build_collectives_url(config.nextcloud_url, config.kollektive_name, page['folder_parts'][:-1], page['folder_parts'][-1])
                    else:
                        from urllib.parse import quote
                        page['url'] = f"{config.nextcloud_url}/apps/collectives/{quote(config.kollektive_name)}"
                else:
                    page['url'] = build_collectives_url(config.nextcloud_url, config.kollektive_name, page['folder_parts'], page['title'])
                ps = status_map.get(page['path']) or status_map.get(page.get('folder', ''))
                if ps:
                    page['status'] = ps.status
                    page['typ'] = ps.typ
                    page['planned_date'] = ps.planned_date.isoformat() if ps.planned_date else ''
                else:
                    page['status'] = ''
                    page['typ'] = ''
                    page['planned_date'] = ''
            return JsonResponse({'success': True, 'pages': pages})
        else:
            return JsonResponse({'success': False, 'message': f'HTTP {r.status_code}'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@login_required
@csrf_exempt
@require_http_methods(["POST"])
def sync_collective_posts(request):
    data = json.loads(request.body)
    collective_pages = data.get('pages', [])
    collective_folders = {p['folder'] for p in collective_pages if p.get('folder')}
    db_folders = set(PageStatus.objects.values_list('path', flat=True))
    to_delete = db_folders - collective_folders
    to_insert = collective_folders - db_folders
    PageStatus.objects.filter(path__in=to_delete).delete()
    for page in collective_pages:
        folder = page.get('folder')
        if folder in to_insert:
            PageStatus.objects.create(path=folder, status='', typ='', planned_date=None)
    return JsonResponse({'success': True, 'message': f'✅ {len(to_delete)} deleted, {len(to_insert)} inserted'})

@login_required
def export_excel(request):
    config = CollectivesConfig.get_config()
    if not config.connected:
        return JsonResponse({'success': False, 'message': 'Not connected'})
    pages_response = get_pages(request)
    pages_dict = json.loads(pages_response.content.decode('utf-8'))
    if not pages_dict.get('success'):
        return JsonResponse({'success': False, 'message': pages_dict.get('message', 'Error')})
    pages = pages_dict.get('pages', [])
    if not pages:
        return JsonResponse({'success': False, 'message': 'No pages found'})
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Collective Pages"
        headers = ['Title', 'Folder / Hierarchy', 'Type', 'Status', 'Planned Date', 'Last Modified', 'Size (KB)', 'Owner', 'Path']
        ws.append(headers)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        status_fills = {
            '📝 In Progress': PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid'),
            '🚀 Ready to Post': PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid'),
            '📤 Posted': PatternFill(start_color='BBDEFB', end_color='BBDEFB', fill_type='solid'),
            '❌ Rejected': PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid'),
        }
        for row_num, page in enumerate(pages, 2):
            modified_date = ''
            if page.get('modified'):
                try:
                    dt = datetime.strptime(page['modified'], '%a, %d %b %Y %H:%M:%S %Z')
                    modified_date = dt.strftime('%d.%m.%Y %H:%M')
                except:
                    modified_date = page['modified']
            size_kb = round(page.get('size', 0) / 1024, 2) if page.get('size') else 0
            status = page.get('status', '')
            typ = page.get('typ', '')
            planned_date = page.get('planned_date', '')
            ws.append([page.get('title', ''), page.get('folder', ''), typ, status, planned_date, modified_date, size_kb, page.get('owner', ''), page.get('path', '')])
            if status and status in status_fills:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_num, column=col).fill = status_fills[status]
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 22
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 25
        ws.column_dimensions['I'].width = 50
        tab_ref = f"A1:I{len(pages) + 1}"
        tab = Table(displayName="CollectivePages", ref=tab_ref)
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        ws.add_table(tab)
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(tmp.name)
        tmp.close()
        with open(tmp.name, 'rb') as f:
            file_content = f.read()
        koll_name = config.kollektive_name.replace(' ', '_').replace('&', 'and') if config.kollektive_name else 'Collective'
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"Collective_Export_{koll_name}_{timestamp}.xlsx"
        response = HttpResponse(file_content, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        import os
        os.unlink(tmp.name)
        return response
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Excel Export Error: {str(e)}'})
